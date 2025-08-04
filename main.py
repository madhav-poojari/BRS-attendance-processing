import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload
from collections import defaultdict

# Google Drive API Setup
SCOPES = ['https://www.googleapis.com/auth/drive']
# secet availble in brs acad gmail

SERVICE_ACCOUNT_FILE = 'brs-fee-processing-ee8420e4d4e9.json'
FOLDER_ID = '1k6ezg3F2KFCPHPl6_gQzUOg7Vv8yhREx'

creds = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES
)
service = build('drive', 'v3', credentials=creds)

from googleapiclient.errors import HttpError
class_date_key = 'Date of class (Coach timezone)'
def download_sheet_as_excel(sheet_id, sheet_name, output_folder):
    """Downloads a Google Sheet as an Excel file."""
    request = service.files().export_media(fileId=sheet_id,
                                            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_path = os.path.join(output_folder, f"{sheet_name}.xlsx")

    with open(file_path, 'wb') as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()

    return file_path
# Get all spreadsheets from the folder
def get_files_from_drive():
    query = f"'{FOLDER_ID}' in parents"
    try:
        results = service.files().list(q=query).execute()
        files = results.get('files', [])
        print(f"Found {len(files)} files in the folder")
        for file in files:
            print(f"  - {file['name']} (ID: {file['id']})")
        return files
    except Exception as e:
        print(f"Error accessing Google Drive: {e}")
        return []

def get_price_per_class(coach_name, student_name, coach_rates):
    """
    Determine the price per class based on the student name and coach rates.

    Parameters:
    - coach_name (str): The normalized name of the coach.
    - student_name (str): The student's name (to check for group classes).
    - coach_rates (dict): Dictionary containing rates for each coach.

    Returns:
    - int/float: The price per class based on the student type.
    """

    # Default pricing structure if coach is not found
    default_prices = {'regular': 0, 'group': 0, 'dual': 0, 'substitution': 0}

    # Get the coach's rates, fallback to default if not found
    rates = coach_rates.get(coach_name, default_prices)

    # Determine class type
    if student_name.lower() == "substitution":
        return rates['substitution']
    elif "," in student_name:  # Group/Dual class if multiple students are present
        return rates['dual']
    else:
        return rates['regular']


# Load coach rates
coach_rates = pd.read_csv('coach_rates.csv')
coach_rates['coach_name'] = coach_rates['coach_name'].str.lower().str.replace(' ', '_')

coach_dict = {}
for _, row in coach_rates.iterrows():
    coach_name = row['coach_name']
    coach_data = row.to_dict()
    coach_dict[coach_name] = coach_data

# User input for month and year
# month_year = input("Enter month and year (MM-YYYY): ")
month_year = "07-2025"
month, year = month_year.split('-')
month, year = int(month), int(year)

# DataFrames for output
master_attendance = []
coach_payout = []

# Color coding
colors = {
    'regular': 'ADD8E6',  # Light Blue
    'substitution': 'B9FF66',  # Orange
    'missing': 'FFFFFF',  # White
    'dual': 'FFFFC5',  # White
    'duplicate': 'FFB6C1'  # Pinkish-red for duplicate/failed processing
}
student_class_count = defaultdict(lambda: defaultdict(int))

# Dictionary to track processed dates for each coach-student pair
processed = defaultdict(lambda: defaultdict(set))

files = get_files_from_drive()
# Process each spreadsheet
for file in files:
    coach_name = file['name'].split('Coach ')[-1].split(' Attendance')[0].lower().replace(' ', '_')
    print(f"processing attendance for {coach_name}")
    sheet_id = file['id']
    sheet_name = file['name']
    file_path = download_sheet_as_excel(sheet_id, sheet_name, 'temp_sheets')
    df = pd.read_excel(file_path)
    df['Coach Name'] = coach_name
    mapping = {}
    for col in df.columns:
        lower = col.lower()
        if 'date' in lower:
            mapping[col] = 'Date of class'
        elif 'name of student' in lower:
            mapping[col] = 'Name of student'
        elif 'substitution' in lower:
            mapping[col] = 'Substitution details'
    df.rename(columns=mapping, inplace=True)
    
    if class_date_key not in df:
        class_date_key = "Date of class"
    df[class_date_key] = pd.to_datetime(df[class_date_key], errors='coerce')
    df = df[(df[class_date_key].dt.month == month) & (df[class_date_key].dt.year == year)]

    for _, row in df.iterrows():
        student_name = row['Name of student']
        class_date = row[class_date_key].date() if pd.notna(row[class_date_key]) else None

        # Determine class type
        if 'substitution' in student_name.lower():
            class_type = 'substitution'
        elif ',' in student_name:
            class_type = 'dual'
        else:
            class_type = 'regular'

        # Check for duplicate processing (same coach, student, and date)
        is_duplicate = False
        if class_date and class_date in processed[coach_name][student_name] and class_type != 'substitution':
            is_duplicate = True

        # Fetch coach rate
        coach_rate = coach_rates[coach_rates['coach_name'] == coach_name]
        if coach_rate.empty:
            fee_processed = False
            color = colors['missing']
            price_per_class = None
        elif is_duplicate:
            fee_processed = False
            color = colors['duplicate']
            price_per_class = coach_rate.iloc[0][class_type]
        else:
            fee_processed = True
            price_per_class = coach_rate.iloc[0][class_type]
            color = colors[class_type]
            # Add date to processed set only if processing is successful
            if class_date:
                processed[coach_name][student_name].add(class_date)

        # Add to master attendance
        master_attendance.append({
            class_date_key: row[class_date_key],
            'Coach Name': coach_name,
            'Student Name': student_name,
            'Class Type': class_type,
            'Fee Processed': fee_processed,
            'Color': color,
            'Substitution': row['Substitution details']
        })
        
        # Only count for payout if fee was processed successfully
        if fee_processed:
            student_class_count[coach_name][student_name] += 1

    print(f"Done with {coach_name}")
for coach_name, students in student_class_count.items():
    for student_name, class_count in students.items():
        price_per_class = get_price_per_class(coach_name, student_name,
                                              coach_dict)  # Function to fetch correct price

        coach_payout.append({
            'Coach Name': coach_name,
            'Student Name': student_name,
            'Price per class': price_per_class,
            'Number of classes': class_count,
            'Total Cost': f'=C{len(coach_payout) + 2}*D{len(coach_payout) + 2}'  # Dynamic Excel formula
        })


# Convert to DataFrame
master_attendance_df = pd.DataFrame(master_attendance)
coach_payout_df = pd.DataFrame(coach_payout)

# Save to Excel
output_file = f'Coach_Payments_{month_year}.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    master_attendance_df.to_excel(writer, sheet_name='Master Attendance', index=False)
    coach_payout_df.to_excel(writer, sheet_name='Coach Payout', index=False)

    # Apply colors to master attendance
    workbook = writer.book
    sheet = workbook['Master Attendance']
    for row_idx, row in enumerate(master_attendance, start=2):
        color_fill = PatternFill(start_color=row['Color'], end_color=row['Color'], fill_type='solid')
        for col_idx in range(1, len(master_attendance_df.columns) + 1):
            sheet.cell(row=row_idx, column=col_idx).fill = color_fill

print(f"Excel file saved as {output_file}")
